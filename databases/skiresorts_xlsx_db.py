'''
    File: skiresorts_xlsx_db.py
    Author: Drew Scott
    Description: Adds all the data from skiresorts.xlsx to summer.skiresorts table.
'''

import openpyxl
import mysql.connector

from passwords import DB, USER, PASSWORD

FIRST_COL = 'skiresort_info_page'
SECOND_COL = 'name'
THIRD_COL = 'website'
FOURTH_COL = 'state'
FIFTH_COL = 'latitude'
SIXTH_COL = 'longitude'

def main():
    # connect to db
    conn = mysql.connector.connect(user=USER, password=PASSWORD, host='localhost', database=DB, auth_plugin='mysql_native_password')

    cursor = conn.cursor()

    # get the worksheet
    workbook = openpyxl.load_workbook(filename="skiresorts.xlsx")
    worksheet = workbook.active

    num_rows = worksheet.max_row
    num_cols = worksheet.max_column

    for row in range(2, num_rows+1):
        # build the query
        query = f"INSERT INTO skiresorts ({FIRST_COL}, {SECOND_COL}, {THIRD_COL}, {FOURTH_COL}, {FIFTH_COL}, {SIXTH_COL}) VALUES ("
        for col in range(1, num_cols+1):
            value = worksheet.cell(row=row, column=col).value
            if value:
                query += '"' + str(value) + '", '
            else:
                query += "NULL, "

        query = query[:-2] + ");"
        print(query)
        cursor.execute(query)

    # commit and close
    cursor.close()
    conn.commit()
    conn.close()

if __name__ == "__main__":
    main()
